from pathlib import Path
import openpyxl

cd = Path.cwd()
filepath = cd/'sample.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb['Sheet1']
# a4 = ws.cell(4,1).value
# c2 = ws.cell(2,3).value
# print('a4 = ',a4,',','c2 = ',c2)
cell_range = ws['A1':'C5']
a4 = cell_range[3][0].value
c2 = cell_range[1][2].value
print('a4 = ',a4,',','c2 = ',c2)